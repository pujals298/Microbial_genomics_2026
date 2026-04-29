"""
Gene Functional Category Classifier — P. aeruginosa PAO1
=========================================================
Classifica gens diferencialment expressats per categories COG.

Fitxer d'entrada per defecte : genes_anotados_def_0.58.xlsx  (full "Genes DEG")
Fitxer de sortida             : genes_anotados_def_0.58_COG_PAO1.xlsx
Fitxer COG local (obligatori) : P.aeruginosa_PAO1_COG  (descarregat de pseudomonas.com)

Ordre de consulta per cada gen:
  0. Fitxer COG local de pseudomonas.com  ← font principal (instantani, oficial)
  1. KEGG (pae:PA####)                    ← per als 88 gens no coberts
  2. UniProt (taxon 208964)               ← fallback API
  3. Keyword inference                    ← últim recurs

Gens amb 2 categories COG: es mostren totes separades per " | "

Caché local (cache_COG_PAO1.json):
  Les consultes API es guarden per evitar repetir-les si l'script s'interromp.

Ús:
  python classify_genes_PAO1.py                        fitxer per defecte
  python classify_genes_PAO1.py altre_fitxer.xlsx      fitxer alternatiu
  python classify_genes_PAO1.py --no-api               només local + keyword
  python classify_genes_PAO1.py --clear-cache          esborra caché API
  python classify_genes_PAO1.py --cog-file ruta/fitxer fitxer COG alternatiu
"""

import os
import sys
import json
import time
import requests
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────
# CONFIGURACIÓ
# ─────────────────────────────────────────────────────────────────

DEFAULT_INPUT   = "genes_anotados_def_0.58.xlsx"
DEFAULT_COG_FILE = "P.aeruginosa_PAO1_COG"
SHEET_NAME      = "Genes DEG"
CACHE_FILE      = "cache_COG_PAO1.json"

ORGANISM_TAXID  = "208964"

# Noms de columnes del fitxer d'entrada
COL_LOCUS   = "Locus Tag"
COL_GENE    = "Nombre del gen"
COL_PRODUCT = "Producto / función"
COL_START   = "Inicio (pb)"
COL_END     = "Fin (pb)"
COL_STRAND  = "Hebra"
COL_LOG2FC  = "Log2FoldChange"
COL_PVAL    = "p-valor ajustado"
COL_EXPR    = "Expresión"

# Categories d'interès
TARGET_CATEGORIES = [
    "Carbohydrate transport and metabolism",
    "Lipid transport and metabolism",
    "Inorganic ion transport and metabolism",
    "Energy production and conversion",
    "Cell wall/membrane/envelope biogenesis",
]

CATEGORY_DISPLAY = {
    "Carbohydrate transport and metabolism":  "Carbohydrates and lipid transport and metabolism",
    "Lipid transport and metabolism":         "Carbohydrates and lipid transport and metabolism",
    "Inorganic ion transport and metabolism": "Inorganic ion transport and metabolism",
    "Energy production and conversion":       "Energy production and conversion",
    "Cell wall/membrane/envelope biogenesis": "Cell wall/membrane/envelope biogenesis",
}

CATEGORY_COLORS = {
    "Carbohydrates and lipid transport and metabolism": "FFF2CC",
    "Inorganic ion transport and metabolism":           "DAE8FC",
    "Energy production and conversion":                 "D5E8D4",
    "Cell wall/membrane/envelope biogenesis":           "F8CECC",
}

# ─────────────────────────────────────────────────────────────────
# DICCIONARI COG COMPLET (nom → lletra)
# ─────────────────────────────────────────────────────────────────

COG_CATEGORIES = {
    "J": "Translation, ribosomal structure and biogenesis",
    "A": "RNA processing and modification",
    "K": "Transcription",
    "L": "Replication, recombination and repair",
    "B": "Chromatin structure and dynamics",
    "D": "Cell cycle control, cell division, chromosome partitioning",
    "Y": "Nuclear structure",
    "V": "Defense mechanisms",
    "T": "Signal transduction mechanisms",
    "M": "Cell wall/membrane/envelope biogenesis",
    "N": "Cell motility",
    "Z": "Cytoskeleton",
    "W": "Extracellular structures",
    "U": "Intracellular trafficking, secretion, and vesicular transport",
    "O": "Posttranslational modification, protein turnover, chaperones",
    "X": "Mobilome: prophages, transposons",
    "C": "Energy production and conversion",
    "G": "Carbohydrate transport and metabolism",
    "E": "Amino acid transport and metabolism",
    "F": "Nucleotide transport and metabolism",
    "H": "Coenzyme transport and metabolism",
    "I": "Lipid transport and metabolism",
    "P": "Inorganic ion transport and metabolism",
    "Q": "Secondary metabolites biosynthesis, transport and catabolism",
    "R": "General function prediction only",
    "S": "Function unknown",
}

# Invers: nom complet → lletra
COG_NAME_TO_LETTER = {v: k for k, v in COG_CATEGORIES.items()}

# ─────────────────────────────────────────────────────────────────
# FONT 0: FITXER COG LOCAL DE PSEUDOMONAS.COM
# ─────────────────────────────────────────────────────────────────

# Diccionari global: {locus_tag: [categoria1, categoria2, ...]}
_local_cog = defaultdict(list)


def load_local_cog_file(filepath):
    """
    Llegeix el fitxer TAB de pseudomonas.com i emplena _local_cog.
    Format esperat (amb capçalera):
      Locus Tag  Product Description  ...  Category
    Gestiona locus tags amb múltiples categories (files duplicades).
    """
    global _local_cog
    if not os.path.isfile(filepath):
        print(f"  AVIS: No es troba el fitxer COG local '{filepath}'")
        print("        Es continuarà amb les fonts API (mes lent).\n")
        return 0

    count = 0
    with open(filepath, "r", encoding="utf-8") as f:
        header = f.readline()  # salta capçalera
        # Detecta índex de columnes
        cols = header.strip().split("\t")
        try:
            lt_col  = cols.index("Locus Tag")
            cat_col = cols.index("Category")
        except ValueError:
            # Fallback: assumeix col 0 = locus tag, col -1 = categoria
            lt_col, cat_col = 0, -1

        for line in f:
            parts = line.strip().split("\t")
            if len(parts) <= max(lt_col, abs(cat_col)):
                continue
            lt  = parts[lt_col].strip()
            cat = parts[cat_col].strip()
            if lt and cat and cat not in _local_cog[lt]:
                _local_cog[lt].append(cat)
                count += 1

    print(f"  Fitxer COG local carregat: {len(_local_cog)} locus tags "
          f"({count} anotacions totals)")
    multi = sum(1 for v in _local_cog.values() if len(v) > 1)
    print(f"  Locus tags amb 2+ categories COG: {multi}\n")
    return len(_local_cog)


def lookup_local_cog(locus_tag):
    """
    Retorna (llista_categories, llista_lletres) per un locus tag,
    o (None, None) si no es troba.
    Gestiona múltiples categories separant-les per " | ".
    """
    cats = _local_cog.get(locus_tag)
    if not cats:
        return None, None

    letters = []
    for cat in cats:
        letter = COG_NAME_TO_LETTER.get(cat)
        if letter:
            letters.append(letter)

    return cats, letters if letters else None

# ─────────────────────────────────────────────────────────────────
# FONT 1: KEGG PAO1 (pae:PA####)
# ─────────────────────────────────────────────────────────────────

KEGG_TO_COG = {
    "carbohydrate metabolism":               "G",
    "lipid metabolism":                      "I",
    "nucleotide metabolism":                 "F",
    "amino acid metabolism":                 "E",
    "metabolism of cofactors and vitamins":  "H",
    "energy metabolism":                     "C",
    "cell motility":                         "N",
    "membrane transport":                    "P",
    "signal transduction":                   "T",
    "replication and repair":                "L",
    "transcription":                         "K",
    "translation":                           "J",
    "folding, sorting and degradation":      "O",
    "cell wall biogenesis":                  "M",
    "biosynthesis of secondary metabolites": "Q",
    "xenobiotics biodegradation":            "Q",
    "drug resistance":                       "V",
    "infectious disease":                    "V",
}


def query_kegg(locus_tag):
    """Consulta KEGG per al gen de PAO1. Retorna lletra COG o None."""
    try:
        r = requests.get(f"https://rest.kegg.jp/get/pae:{locus_tag}", timeout=10)
        if r.status_code != 200:
            return None
        text = r.text.lower()
        for key, letter in KEGG_TO_COG.items():
            if key in text:
                return letter
        for letter in COG_CATEGORIES:
            if f"cog{letter.lower()}" in text or f"[{letter.lower()}]" in text:
                return letter
        return None
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────
# FONT 2: UNIPROT (taxon 208964)
# ─────────────────────────────────────────────────────────────────

def query_uniprot(gene_name, locus_tag):
    """Consulta UniProt filtrant per PAO1. Retorna lletra COG o None."""
    try:
        queries = []
        if locus_tag and locus_tag not in ["", "nan", "-"]:
            queries.append(f"gene:{locus_tag} AND organism_id:{ORGANISM_TAXID}")
        if gene_name and gene_name not in ["", "nan", "-", "—"]:
            queries.append(f"gene:{gene_name} AND organism_id:{ORGANISM_TAXID}")

        for query in queries:
            r = requests.get(
                "https://rest.uniprot.org/uniprotkb/search",
                params={"query": query,
                        "fields": "gene_names,xref_cog,go_p,go_f",
                        "format": "json", "size": 1},
                timeout=10
            )
            if r.status_code != 200:
                continue
            results = r.json().get("results", [])
            if not results:
                continue
            entry = results[0]
            if str(entry.get("organism", {}).get("taxonId", "")) != ORGANISM_TAXID:
                continue
            for xref in entry.get("uniProtKBCrossReferences", []):
                if xref.get("database") == "COG":
                    for prop in xref.get("properties", []):
                        if prop.get("key") == "FunctionalCategory":
                            val = prop.get("value", "").strip().upper()
                            if val in COG_CATEGORIES:
                                return val
        return None
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────
# FONT 3: KEYWORD INFERENCE (fallback)
# ─────────────────────────────────────────────────────────────────

def infer_cog_from_product(product_text):
    """Infereix lletra COG per paraules clau del producte."""
    if not product_text or str(product_text).strip() in ["", "nan", "-", "—"]:
        return "S"

    text = str(product_text).lower()
    rules = [
        (["carbohydrate", "sugar", "glyco", "glucos", "fructos", "lactose", "maltose",
          "sucrose", "galactos", "xylose", "polysaccharide", "glycolysis",
          "gluconeogenesis", "phosphotransferase system", "pts "], "G"),
        (["lipid", "fatty acid", "acyl", "phospholipid", "membrane lipid",
          "triglyceride", "beta-oxidation", "lipase", "acyl-coa"], "I"),
        (["inorganic ion", "phosphate transport", "sulfate transport", "iron transport",
          "metal ion", "zinc", "copper transport", "potassium transport",
          "abc transporter", "atp-binding cassette", "magnesium", "manganese",
          "cobalt", "ferric", "ferrous", "molybdate", "sulfur", "iron-sulfur",
          "siderophore", "pyoverdine", "pyochelin"], "P"),
        (["atp synthase", "electron transport", "oxidative phosphorylation",
          "respiratory", "cytochrome", "nadh dehydrogenase", "succinate dehydrogenase",
          "fumarate reductase", "energy production", "atpase", "proton pump",
          "quinol", "ubiquinol", "menaquinone", "ferredoxin", "hydrogenase",
          "nitrogenase", "nitric-oxide reductase", "nitrite reductase",
          "nitrous oxide", "azurin", "nitrate reductase"], "C"),
        (["cell wall", "peptidoglycan", "murein", "outer membrane",
          "lipopolysaccharide", "lps", "capsule", "envelope", "omp", "porin",
          "penicillin-binding", "transpeptidase", "transglycosylase",
          "alginate", "rhamnolipid"], "M"),
        (["ribosom", "translation", "trna", "aminoacyl", "initiation factor",
          "elongation factor", "termination factor", "rrna"], "J"),
        (["transcription", "rna polymerase", "sigma factor", "transcriptional regulator",
          "dna-binding", "helix-turn-helix", "luxr", "laci"], "K"),
        (["dna repair", "dna replication", "recombinase", "transposase",
          "integrase", "helicase", "topoisomerase", "gyrase", "nuclease",
          "exonuclease", "dna ligase"], "L"),
        (["amino acid", "glutamate", "glutamine", "aspartate", "asparagine",
          "serine", "threonine", "cysteine", "methionine", "lysine", "arginine",
          "ornithine", "proline", "tryptophan", "phenylalanine", "tyrosine",
          "histidine", "valine", "leucine", "isoleucine"], "E"),
        (["coenzyme", "cofactor", "biotin", "folate", "riboflavin", "thiamine",
          "pantothenate", "lipoate", "cobalamin", "vitamin", "nad biosynthesis",
          "fad", "heme biosynthesis", "porphyrin"], "H"),
        (["signal transduction", "two-component", "sensor histidine kinase",
          "response regulator", "chemotaxis", "di-gmp", "diguanylate cyclase",
          "phosphodiesterase"], "T"),
        (["flagell", "motility", "fli", "flh", "flg", "mot", "fla",
          "twitching", "type iv pili", "pilb", "pilc", "pild"], "N"),
        (["type iii secretion", "type iv secretion", "type vi secretion",
          "type ii secretion", "sec translocon", "tat pathway", "twin-arginine",
          "xcp", "hsi", "vgr"], "U"),
        (["defense", "crispr", "cas protein", "bacteriocin",
          "antibiotic resistance", "efflux pump", "mex", "opr"], "V"),
        (["chaperone", "heat shock protein", "hsp", "groel", "groes", "dnak",
          "dnaj", "grpe", "lon protease", "clp protease", "htpg"], "O"),
        (["nucleotide", "purine", "pyrimidine", "guanosine", "adenosine",
          "thymidine", "uridine", "cytidine", "phosphoribosyl"], "F"),
        (["quorum sensing", "las", "rhl", "pqs", "acyl-homoserine",
          "phenazine", "pyocyanin", "elastase", "alkaline protease",
          "exotoxin", "rhamnolipid biosynthesis"], "Q"),
        (["transposon", "insertion sequence", "is element", "prophage",
          "phage", "integron", "plasmid"], "X"),
        (["hypothetical protein", "unknown function", "uncharacterized",
          "conserved hypothetical"], "S"),
        (["predicted", "possible", "probable", "putative"], "R"),
    ]
    for keywords, letter in rules:
        if any(kw in text for kw in keywords):
            return letter
    return "S"

# ─────────────────────────────────────────────────────────────────
# CACHÉ LOCAL (per a les consultes API)
# ─────────────────────────────────────────────────────────────────

def load_cache(path):
    if os.path.isfile(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

# ─────────────────────────────────────────────────────────────────
# ORQUESTRADOR PRINCIPAL
# ─────────────────────────────────────────────────────────────────

def get_cog_category(locus_tag, gene_name, product, cache, use_api=True):
    """
    Consulta les fonts en ordre:
      0. Fitxer local pseudomonas.com  (instantani, oficial)
      1. KEGG pae                      (API)
      2. UniProt taxon 208964          (API)
      3. Keyword inference             (fallback)
    """
    # Font 0: fitxer local pseudomonas.com
    cats, letters = lookup_local_cog(locus_tag)
    if cats:
        cat_str    = " | ".join(cats)
        letter_str = " | ".join(letters) if letters else "S"
        return letter_str, cat_str, "pseudomonas.com (local)"

    # Caché de consultes API prèvies
    cache_key = locus_tag or gene_name
    if cache_key in cache:
        e = cache[cache_key]
        return e["letter"], e["category"], e["source"] + " [caché]"

    if use_api:
        # Font 1: KEGG
        letter = query_kegg(locus_tag)
        if letter:
            cat = COG_CATEGORIES.get(letter, "Function unknown")
            cache[cache_key] = {"letter": letter, "category": cat, "source": "KEGG (pae)"}
            return letter, cat, "KEGG (pae)"
        time.sleep(0.2)

        # Font 2: UniProt
        letter = query_uniprot(gene_name, locus_tag)
        if letter:
            cat = COG_CATEGORIES.get(letter, "Function unknown")
            cache[cache_key] = {"letter": letter, "category": cat, "source": "UniProt"}
            return letter, cat, "UniProt"
        time.sleep(0.2)

    # Font 3: keyword inference
    letter = infer_cog_from_product(product)
    cat    = COG_CATEGORIES.get(letter, "Function unknown")
    cache[cache_key] = {"letter": letter, "category": cat, "source": "Keyword inference"}
    return letter, cat, "Keyword inference"

# ─────────────────────────────────────────────────────────────────
# NOM DEL FITXER DE SORTIDA
# ─────────────────────────────────────────────────────────────────

def build_output_filename(input_file):
    base, ext = os.path.splitext(input_file)
    return f"{base}_COG_PAO1{ext}"

# ─────────────────────────────────────────────────────────────────
# PROCESSAMENT PRINCIPAL
# ─────────────────────────────────────────────────────────────────

def process_genes(input_file, output_file, cog_file, use_api=True):
    print(f"\nLlegint fitxer d'entrada: {input_file}")
    df = pd.read_excel(input_file, sheet_name=SHEET_NAME)
    print(f"  Gens carregats: {len(df)}")

    print(f"\nCarregant fitxer COG local: {cog_file}")
    load_local_cog_file(cog_file)

    cache = load_cache(CACHE_FILE)
    if cache:
        print(f"Caché API carregada: {len(cache)} entrades\n")

    results      = []
    source_counts = {}

    print("Classificant gens...")
    print(f"  {'#':<5} {'Locus Tag':<10} {'Gen':<12} {'Lletra':<8} "
          f"{'Categoria':<45} {'Font'}")
    print("  " + "-" * 100)

    for i, row in df.iterrows():
        locus   = str(row.get(COL_LOCUS,   "")).strip()
        gene    = str(row.get(COL_GENE,    "")).strip()
        product = str(row.get(COL_PRODUCT, "")).strip()
        start   = str(row.get(COL_START,   "")).strip()
        end     = str(row.get(COL_END,     "")).strip()
        strand  = str(row.get(COL_STRAND,  "")).strip()
        log2fc  = row.get(COL_LOG2FC, "")
        pval    = row.get(COL_PVAL,   "")
        expr    = str(row.get(COL_EXPR,    "")).strip()

        letter_str, cat_str, source = get_cog_category(
            locus, gene, product, cache, use_api=use_api
        )

        gene_display = gene if gene not in ["", "nan", "—"] else "—"
        # Per mostrar al terminal, trunca si hi ha múltiples categories
        cat_display = cat_str if len(cat_str) <= 44 else cat_str[:41] + "..."
        print(f"  [{i+1:>3}/{len(df)}] {locus:<10} {gene_display:<12} "
              f"{letter_str:<8} {cat_display:<45} {source}")

        src_clean = source.replace(" [caché]", "")
        source_counts[src_clean] = source_counts.get(src_clean, 0) + 1

        results.append({
            COL_LOCUS:        locus,
            COL_GENE:         gene,
            COL_PRODUCT:      product,
            COL_START:        start,
            COL_END:          end,
            COL_STRAND:       strand,
            COL_LOG2FC:       log2fc,
            COL_PVAL:         pval,
            COL_EXPR:         expr,
            "COG Letter":     letter_str,
            "COG Category":   cat_str,
            "Source":         source,
        })

        # Desa caché cada 10 gens (per si s'interromp)
        if (i + 1) % 10 == 0:
            save_cache(cache, CACHE_FILE)

    save_cache(cache, CACHE_FILE)

    df_all = pd.DataFrame(results)

    # Per filtrar les categories d'interès cal tenir en compte que
    # un gen pot tenir 2 categories separades per " | "
    def is_target(cat_str):
        return any(tc in cat_str for tc in TARGET_CATEGORIES)

    df_filtered = df_all[df_all["COG Category"].apply(is_target)].copy()

    def get_display_category(cat_str):
        for tc in TARGET_CATEGORIES:
            if tc in cat_str:
                return CATEGORY_DISPLAY[tc]
        return ""

    df_filtered["Target Category"] = df_filtered["COG Category"].apply(get_display_category)

    # Resum
    total = len(df_all)
    print("\n" + "=" * 65)
    print("RESUM")
    print(f"  Total gens classificats:           {total}")
    print(f"  Gens en categories d'interes:      {len(df_filtered)}")
    print(f"\n  FONTS UTILITZADES:")
    for src, n in sorted(source_counts.items(), key=lambda x: -x[1]):
        pct = n / total * 100
        print(f"    {src:<35} {n:>4} gens ({pct:.1f}%)")
    unknown = df_all[df_all["COG Category"] == "Function unknown"]
    print(f"\n  Gens sense categoria (S):          "
          f"{len(unknown)} ({len(unknown)/total*100:.1f}%)")
    multi = df_all[df_all["COG Letter"].str.contains(r"\|", na=False)]
    print(f"  Gens amb 2+ categories COG:        {len(multi)}")
    print("=" * 65)

    write_excel(df_all, df_filtered, output_file)
    print(f"\nFitxer generat : {output_file}")
    print(f"Caché guardada : {CACHE_FILE}")

# ─────────────────────────────────────────────────────────────────
# ESCRIPTURA EXCEL
# ─────────────────────────────────────────────────────────────────

def write_excel(df_all, df_filtered, output_file):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_all.to_excel(writer,
                        sheet_name="All Genes - COG Classification",
                        index=False)
        df_filtered.to_excel(writer,
                             sheet_name="Target Categories",
                             index=False)
    wb = load_workbook(output_file)
    _format_sheet(wb["All Genes - COG Classification"], df_all,     filtered=False)
    _format_sheet(wb["Target Categories"],              df_filtered, filtered=True)
    wb.save(output_file)


# Estils
HEADER_FILL  = PatternFill("solid", start_color="2F4F8F")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BODY_FONT    = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
UP_FILL   = PatternFill("solid", start_color="C6EFCE")  # verd: sobreexpressat
DOWN_FILL = PatternFill("solid", start_color="FFC7CE")  # vermell: reprimit

# Color per font (pestanya All Genes)
SOURCE_COLORS = {
    "pseudomonas.com (local)": "E2EFDA",  # verd clar
    "KEGG (pae)":              "DDEBF7",  # blau clar
    "UniProt":                 "EEE2EF",  # violeta clar
    "Keyword inference":       "F2F2F2",  # gris clar
}


def _format_sheet(sheet, df, filtered=False):
    # Capçalera
    for cell in sheet[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER
    sheet.row_dimensions[1].height = 30

    cols     = list(df.columns)
    expr_idx = next((i+1 for i, c in enumerate(cols) if "expresi" in c.lower()), None)
    src_idx  = next((i+1 for i, c in enumerate(cols) if c == "Source"), None)
    tcat_idx = next((i+1 for i, c in enumerate(cols) if "target" in c.lower()), None)

    for row in sheet.iter_rows(min_row=2):
        expr_val = str(row[expr_idx - 1].value or "").lower() if expr_idx else ""
        src_val  = str(row[src_idx  - 1].value or "")        if src_idx  else ""
        cat_val  = str(row[tcat_idx - 1].value or "")        if tcat_idx else ""

        if filtered:
            # Color per categoria
            bg = None
            for cat_key, color in CATEGORY_COLORS.items():
                if cat_key in cat_val:
                    bg = PatternFill("solid", start_color=color)
                    break
        else:
            # Color per font
            src_clean = src_val.replace(" [caché]", "")
            color = SOURCE_COLORS.get(src_clean)
            bg = PatternFill("solid", start_color=color) if color else None

        for cell in row:
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if bg:
                cell.fill = bg
            elif "sobre" in expr_val or "over" in expr_val or "up" in expr_val:
                cell.fill = UP_FILL
            elif "reprim" in expr_val or "down" in expr_val:
                cell.fill = DOWN_FILL

    # Amplades
    widths = {"A": 12, "B": 16, "C": 44, "D": 12, "E": 12,
              "F": 8,  "G": 16, "H": 18, "I": 16, "J": 12,
              "K": 44, "L": 22}
    if filtered:
        widths["M"] = 46
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

# ─────────────────────────────────────────────────────────────────
# PUNT D'ENTRADA
# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    args = sys.argv[1:]

    clear_cache = "--clear-cache" in args
    use_api     = "--no-api" not in args
    args_flags  = [a for a in args if a.startswith("--")]
    args_pos    = [a for a in args if not a.startswith("--")]

    # --cog-file ruta/fitxer
    cog_file = DEFAULT_COG_FILE
    if "--cog-file" in args_flags:
        idx = args.index("--cog-file")
        if idx + 1 < len(args):
            cog_file = args[idx + 1]
            args_pos = [a for a in args_pos if a != cog_file]

    input_file = args_pos[0] if args_pos else DEFAULT_INPUT

    if not os.path.isfile(input_file):
        print(f"\nERROR: No es troba el fitxer d'entrada '{input_file}'")
        print(f"""
Us:
  python classify_genes_PAO1.py                          fitxer per defecte
  python classify_genes_PAO1.py altre.xlsx               fitxer alternatiu
  python classify_genes_PAO1.py --no-api                 nomes local + keyword
  python classify_genes_PAO1.py --clear-cache            esborra cache API
  python classify_genes_PAO1.py --cog-file ruta/fitxer   fitxer COG alternatiu

Fitxers necessaris a la mateixa carpeta:
  - {DEFAULT_INPUT}
  - {DEFAULT_COG_FILE}   (descarregat de pseudomonas.com)
""")
        sys.exit(1)

    if clear_cache and os.path.isfile(CACHE_FILE):
        os.remove(CACHE_FILE)
        print(f"Cache esborrada: {CACHE_FILE}")

    output_file = build_output_filename(input_file)
    process_genes(input_file, output_file, cog_file, use_api=use_api)
