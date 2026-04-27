"""
Descripción: Cruza los resultados de DESeq2 (locus tags) con un genoma anotado
             en formato GenBank para generar una tabla Excel con la información
             de cada gen diferencialmente expresado.

Requisitos (instalar una sola vez con pip):
    pip install biopython pandas openpyxl

Uso:
    1. Pon este script en la misma carpeta que tus archivos
    2. Cambia las rutas en la sección "CONFIGURACIÓN" más abajo
    3. Ejecuta: python anotar_genes_deseq2.py
"""

# --- LIBRERÍAS ---
# Biopython: para leer archivos GenBank
# Pandas: para manejar tablas de datos
# Openpyxl: para dar formato al Excel de salida

from Bio import SeqIO # SeqIO es el módulo que sabe leer archivos de secuencias: GenBank, FASTA, etc. Sin esto, Python no sabría qué hacer con el .gbk
import pandas as pd # La librería estándar para manejar tablas de datos en Python
from openpyxl import load_workbook # Se encarga de crear y dar formato a archivos .xlsx. Con pandas podemos escribir los datos, pero para cosas de formato se necesita este
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURACIÓN 
# =============================================================================

ARCHIVO_DESEQ2  = "results_deseq2.xlsx"   # Tu output de DESeq2
ARCHIVO_GENBANK = "Pseudomonas_aeruginosa_PAO1_107.gbk" # Tu genoma de referencia anotado
ARCHIVO_SALIDA  = "genes_anotados_def.xlsx"      # Nombre del Excel que se generará

# Nombre exacto de la columna con los locus tags en tu archivo de DESeq2
COLUMNA_LOCUS_TAG = "GeneID"

# Filtros de significancia (puedes ajustarlos)
PADJ_UMBRAL     = 0.05   # p-valor ajustado máximo (FDR)
LOG2FC_UMBRAL   = 0.58    # Log2FoldChange mínimo en valor absoluto
                         # (0.58 = el mas adecuado segun literatura, captura flujos metabolicos y regulacion)

# =============================================================================

# Definimos la función "leer_deseq2". Es un bloque de instrucciones con nombre que puedes ejecutar cuando quieras

def leer_deseq2(ruta): 
    """Lee el Excel de DESeq2 y devuelve solo los genes significativos."""
    print(f"[1/4] Leyendo resultados de DESeq2: {ruta}")
    
        # Usamos pandas (pd) para leer el excel y lo guardamos en una función "dataframe" (df).

    df = pd.read_excel(ruta, dtype={COLUMNA_LOCUS_TAG: str})
    print(f"      → {len(df)} genes en total")

        # Filtramos por significancia estadística
        # Los corchetes en pandas funcionan como un filtro: df[condición] devuelve solo las filas que cumplen esa condición
        # & significa "AND". Las 2 condiciones deben cumplirse a la vez. El .abs significa que toma solo valores absolutos

    df_sig = df[
        (df["padj"] < PADJ_UMBRAL) &
        (df["log2FoldChange"].abs() >= LOG2FC_UMBRAL)
    ].copy()

    print(f"      → {len(df_sig)} genes significativos "
          f"(padj < {PADJ_UMBRAL} y |log2FC| >= {LOG2FC_UMBRAL})")
    return df_sig # Devolvemos el dataframe filtrado con los genes significativos. Esta tabla filtrada es lo que sale de la función para que el resto del script pueda usarla


def construir_diccionario_genbank(ruta):
    """
    Lee el GenBank y construye un diccionario donde la clave es el locus_tag
    y el valor es un diccionario con la info del gen.
    """
    print(f"[2/4] Leyendo genoma GenBank: {ruta}")
    
    anotacion = {} # Comenzamos con un diccionario vacío

    # SeqIO.parse lee cada registro del archivo .gbk (en PAO1 hay uno solo: el cromosoma)
    for registro in SeqIO.parse(ruta, "genbank"):
        for feature in registro.features:
            # Solo nos interesan los CDS (regiones codificantes de proteínas)
            if feature.type != "CDS": # "Si esta feature NO es un CDS, continúa"
                continue
            
            # .get() devuelve una lista o [] si no existe ese campo
            # El .get("locus_tag", [None]) intenta obtener el locus tag, y si no existe devuelve [None] en vez de dar error. 
            # El [0] coge el primer elemento de esa lista. Si al final locus está vacío, lo saltamos.
            locus = feature.qualifiers.get("locus_tag", [None])[0]
            if locus is None:
                continue
            # Básicamente la estructura de feature.qualifiers.get es ("Lo que tienes que buscar", [Lo que tienes que poner si no existe]) 

            # Para cada CDS que pase los filtros, le creamos un compartimento en el diccionario "anotacion" con la etiqueta del locus tag

            anotacion[locus] = {
                "gene":      feature.qualifiers.get("gene", ["—"])[0],
                "producto":  feature.qualifiers.get("product", ["hypothetical"])[0],
                "inicio":    int(feature.location.start) + 1,  # +1: GenBank es 0-based
                "fin":       int(feature.location.end),
                "hebra":     "+" if feature.location.strand == 1 else "-",
            }

    print(f"      → {len(anotacion)} CDS encontrados en el genoma")
    return anotacion

# Definimos la función de cruzar los datos, que recibe los datos de los dos parámetros anteriores que hemos definido

def cruzar_datos(df_sig, anotacion):
    """Une los resultados de DESeq2 con la anotación del GenBank."""
    print("[3/4] Cruzando locus tags con la anotación...")

    # Creamos dos listas. A diferencia de los diccionarios no tienen etiquetas, simplemente guardan elementos en orden

    filas = []
    no_encontrados = []

    for _, fila in df_sig.iterrows(): # .iterrows() es una función de pandas que recorre la tabla fila a fila. 
        locus = fila[COLUMNA_LOCUS_TAG]
        info  = anotacion.get(locus)

        #Si .get() no encontró el locus tag, info valdrá None, y entonces lo añadimos a la lista no_encontrados con .append()

        if info is None:
            no_encontrados.append(locus)
            continue

        # Clasificamos si el gen está sobre o infraexpresado. Solo miramos el símbolo del fold change: positivo significa que el gen se expresa más en el mutante que en el WT, negativo lo contrario
        lfc = fila["log2FoldChange"]
        if lfc > 0:
            expresion = "Sobreexpresado"
        else:
            expresion = "Reprimido"

        filas.append({ # La función info la hemos creado antes, cuando hicimos anotacion.get(locus). Esto significa que
            # estamos buscando ese locus tag en el diccionario del GenBank que construimos en la sección anterior. Si lo encuentra, info se convierte en el diccionario pequeño de ese gen
            "Locus Tag":          locus,
            "Nombre del gen":     info["gene"],
            "Producto / función": info["producto"],
            "Inicio (pb)":        info["inicio"],
            "Fin (pb)":           info["fin"],
            "Hebra":              info["hebra"],
            "Log2FoldChange":     round(lfc, 4),
            "p-valor ajustado":   fila["padj"],
            "Expresión":          expresion,
        })

# Por cada gen creamos un diccionario pequeño con toda su informacion. Al final del bucle, "filas" será una lista de diccionarios, uno por gen.

    if no_encontrados:
        print(f"      ⚠ {len(no_encontrados)} locus tags no encontrados en el GenBank:")
        for lt in no_encontrados:
            print(f"        - {lt}")

    df_out = pd.DataFrame(filas) # pd.DataFrame(filas) convierte la lista de diccionarios en una tabla de pandas, con las columnas basadas en las claves
    # Ordenamos: primero sobreexpresados, luego infraexpresados, por |log2FC|
    df_out["abs_lfc"] = df_out["Log2FoldChange"].abs() # Todo esto es para crear una "columna temporal" que nos ordenará los valores de l2Fc
    df_out = df_out.sort_values( 
        ["Expresión", "abs_lfc"], ascending=[True, False]
    ).drop(columns="abs_lfc") # Después de ordenar, eliminamos la columna temporal con .drop()

    return df_out

# Esta función es la más larga pero conceptualmente la más sencilla, es para dar formato visual al Excel

def exportar_excel(df, ruta_salida): 
    """Escribe el DataFrame a Excel con formato visual claro."""
    print(f"[4/4] Generando Excel: {ruta_salida}")

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer: # pd.ExcelWriter crea el archivo Excel
        df.to_excel(writer, index=False, sheet_name="Genes DEG") # .to_excel() vuelca la tabla dentro

    # --- Formato con openpyxl ---
    # Aquí es donde entra openpyxl. Cargamos el archivo recién creado en una variable wb (workbook), y seleccionamos la hoja concreta en ws (worksheet)
    wb = load_workbook(ruta_salida)
    ws = wb["Genes DEG"]

    # Colores
    COLOR_CABECERA    = "2C3E50"   # Azul oscuro
    COLOR_SOBRE       = "E8F5E9"   # Verde claro
    COLOR_INFRA       = "FDECEA"   # Rojo claro
    COLOR_TEXTO_BLANCO = "FFFFFF"

    # Cabeceras
    for celda in ws[1]:
        celda.font      = Font(bold=True, color=COLOR_TEXTO_BLANCO, name="Arial", size=11)
        celda.fill      = PatternFill("solid", fgColor=COLOR_CABECERA)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# Font, PatternFill y Alignment son las herramientas que importamos de openpyxl al principio del script

    ws.row_dimensions[1].height = 32

    # Filas de datos
    col_expresion = [c.column for c in ws[1] if c.value == "Expresión"][0]

# Bucle dentro de bucle: El bucle exterior recorre cada fila, el interior recorre cada celda de esa fila
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row):
        expresion = fila[col_expresion - 1].value
        color = COLOR_SOBRE if expresion == "Sobreexpresado" else COLOR_INFRA

        for celda in fila:
            celda.font      = Font(name="Arial", size=10)
            celda.fill      = PatternFill("solid", fgColor=color)
            celda.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuste automático del ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10) #
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50) # Ajusta automáticamente el ancho de cada columna según el contenido más largo que tenga

    # Congelamos la primera fila para facilitar la navegación
    ws.freeze_panes = "A2"

    # Añadimos una hoja de resumen
    ws_res = wb.create_sheet("Resumen")
    sobre   = df[df["Expresión"] == "Sobreexpresado"]
    infra   = df[df["Expresión"] == "Reprimido"]

    resumen = [
        ["Resumen del análisis DEG — Pseudomonas aeruginosa PAO1 (WT vs ΔpurM)"],
        [],
        ["Parámetro",             "Valor"],
        ["Total genes analizados", len(df)],
        ["Genes sobreexpresados",  len(sobre)],
        ["Genes reprimidos",       len(infra)],
        ["Umbral p-adj",           PADJ_UMBRAL],
        ["Umbral |log2FC|",        LOG2FC_UMBRAL],
    ]

    for i, fila_res in enumerate(resumen, start=1):
        for j, val in enumerate(fila_res, start=1):
            celda = ws_res.cell(row=i, column=j, value=val)
            if i == 1 or (i == 3):
                celda.font = Font(bold=True, name="Arial", size=11)

    ws_res.column_dimensions["A"].width = 32
    ws_res.column_dimensions["B"].width = 20

    wb.save(ruta_salida)

# =============================================================================
# EJECUCIÓN PRINCIPAL
# =============================================================================
# AQUI es donde realmente se ejecutan las funciones que hemos estado definiendo en orden. Se van pasando el resultado de cada una a la siguiente.
if __name__ == "__main__":
    df_sig   = leer_deseq2(ARCHIVO_DESEQ2)
    anotacion = construir_diccionario_genbank(ARCHIVO_GENBANK)
    df_final  = cruzar_datos(df_sig, anotacion)
    exportar_excel(df_final, ARCHIVO_SALIDA)
