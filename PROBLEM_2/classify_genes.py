"""
Gene Functional Category Classifier
====================================
Classifica gens diferencialment expressats per categories COG
a partir d'un Excel d'entrada i genera un Excel de sortida amb:
  - Sheet 1: Tots els gens amb la seva categoria COG
  - Sheet 2: Gens filtrats per les 4 categories d'interès
"""

import pandas as pd
import requests
import time
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURACIÓ: modifica aquí si cal
# ─────────────────────────────────────────────

# Nom de les columnes del teu Excel d'entrada
# Ajusta'ls als noms reals de les teves columnes
COL_LOCUS_TAG   = "Locus Tag"       # columna amb el locus tag (ex: RSP_0001)
COL_GENE_NAME   = "Gene Name"       # columna amb el nom del gen
COL_PRODUCT     = "Product"         # columna amb el producte/funció
COL_LOCATION    = "Location"        # columna amb la ubicació en el genoma
COL_STRAND      = "Strand"          # columna amb la cadena (+/-)
COL_REGULATION  = "Regulation"      # columna amb sobreexpresat/reprimit

# Les 4 categories d'interès (noms exactes COG)
TARGET_CATEGORIES = [
    "Carbohydrate transport and metabolism",
    "Lipid transport and metabolism",
    "Inorganic ion transport and metabolism",
    "Energy production and conversion",
    "Cell wall/membrane/envelope biogenesis",
]

# Categoria combinada que es mostrarà a la taula filtrada
CATEGORY_DISPLAY = {
    "Carbohydrate transport and metabolism":    "Carbohydrates and lipid transport and metabolism",
    "Lipid transport and metabolism":           "Carbohydrates and lipid transport and metabolism",
    "Inorganic ion transport and metabolism":   "Inorganic ion transport and metabolism",
    "Energy production and conversion":         "Energy production and conversion",
    "Cell wall/membrane/envelope biogenesis":   "Cell wall/membrane/envelope biogenesis",
}

# Colors per a cada categoria (fons de cel·la, hex sense #)
CATEGORY_COLORS = {
    "Carbohydrates and lipid transport and metabolism": "FFF2CC",  # groc clar
    "Inorganic ion transport and metabolism":           "DAE8FC",  # blau clar
    "Energy production and conversion":                 "D5E8D4",  # verd clar
    "Cell wall/membrane/envelope biogenesis":           "F8CECC",  # rosa clar
}

# ─────────────────────────────────────────────
# DICCIONARI DE CATEGORIES COG
# ─────────────────────────────────────────────

COG_CATEGORIES = {
    "J": "Translation, ribosomal structure and biogenesis",
    "A": "RNA processing and modification",
    "K": "Transcription",
    "L": "Replication, recombination and repair",
    "B": "Chromatin structure and dynamics",
    "D": "Cell cycle control, cell division, chromosome partitioning",
    "Y": "Nuclear structure",
    "V": "Defense mechanisms",
    "T": "Signal transduction mechanisms",
    "M": "Cell wall/membrane/envelope biogenesis",
    "N": "Cell motility",
    "Z": "Cytoskeleton",
    "W": "Extracellular structures",
    "U": "Intracellular trafficking, secretion, and vesicular transport",
    "O": "Posttranslational modification, protein turnover, chaperones",
    "X": "Mobilome: prophages, transposons",
    "C": "Energy production and conversion",
    "G": "Carbohydrate transport and metabolism",
    "E": "Amino acid transport and metabolism",
    "F": "Nucleotide transport and metabolism",
    "H": "Coenzyme transport and metabolism",
    "I": "Lipid transport and metabolism",
    "P": "Inorganic ion transport and metabolism",
    "Q": "Secondary metabolites biosynthesis, transport and catabolism",
    "R": "General function prediction only",
    "S": "Function unknown",
}

# ─────────────────────────────────────────────
# FUNCIONS DE CONSULTA A BASES DE DADES
# ─────────────────────────────────────────────

def query_eggnog_by_locus(locus_tag):
    """
    Intenta obtenir la categoria COG via NCBI Gene per locus tag.
    Retorna la lletra COG o None.
    """
    try:
        # Cerca a NCBI Gene per locus tag
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        params = {
            "db": "gene",
            "term": f"{locus_tag}[Gene Name] OR {locus_tag}[Locus Tag]",
            "retmode": "json",
            "retmax": 1
        }
        r = requests.get(search_url, params=params, timeout=10)
        data = r.json()
        ids = data.get("esearchresult", {}).get("idlist", [])
        if not ids:
            return None

        # Obtenir informació del gen
        fetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
        params2 = {"db": "gene", "id": ids[0], "rettype": "gene_table", "retmode": "text"}
        r2 = requests.get(fetch_url, params=params2, timeout=10)
        # Busca si hi ha informació COG en el text retornat
        text = r2.text
        for cog_letter in COG_CATEGORIES:
            if f"COG{cog_letter}" in text or f"[{cog_letter}]" in text:
                return cog_letter
        return None
    except Exception:
        return None


def query_uniprot_cog(gene_name, locus_tag):
    """
    Consulta UniProt per nom de gen o locus tag i intenta extreure la categoria COG.
    """
    try:
        queries = []
        if gene_name and str(gene_name).strip() not in ["", "nan", "-"]:
            queries.append(f"gene:{gene_name}")
        if locus_tag and str(locus_tag).strip() not in ["", "nan", "-"]:
            queries.append(f"xref:ncbigi-{locus_tag}")

        for query in queries:
            url = "https://rest.uniprot.org/uniprotkb/search"
            params = {
                "query": query,
                "fields": "gene_names,protein_name,go,cc_function,xref_cog",
                "format": "json",
                "size": 1
            }
            r = requests.get(url, params=params, timeout=10)
            if r.status_code != 200:
                continue
            data = r.json()
            results = data.get("results", [])
            if not results:
                continue

            # Cerca referència COG en els cross-references
            entry = results[0]
            xrefs = entry.get("uniProtKBCrossReferences", [])
            for xref in xrefs:
                if xref.get("database") == "COG":
                    props = xref.get("properties", [])
                    for prop in props:
                        if prop.get("key") == "FunctionalCategory":
                            return prop.get("value", "").strip()
            # Si no hi ha COG explícit, intenta inferir per GO terms
            go_terms = []
            for xref in xrefs:
                if xref.get("database") == "GO":
                    props = xref.get("properties", [])
                    for prop in props:
                        if prop.get("key") == "GoTerm":
                            go_terms.append(prop.get("value", ""))
            if go_terms:
                return infer_cog_from_go(go_terms)
        return None
    except Exception:
        return None


def infer_cog_from_go(go_terms):
    """
    Infereix la lletra COG a partir de termes GO.
    Mapeig simplificat per als casos més comuns.
    """
    go_text = " ".join(go_terms).lower()
    mappings = [
        (["carbohydrate transport", "carbohydrate metabolic", "sugar transport", "glycolysis",
          "gluconeogenesis", "polysaccharide"], "G"),
        (["lipid transport", "lipid metabolic", "fatty acid", "membrane lipid"], "I"),
        (["inorganic ion", "metal ion transport", "phosphate transport",
          "iron transport", "zinc", "copper", "sulfur"], "P"),
        (["atp synthesis", "electron transport", "oxidative phosphorylation",
          "energy derivation", "cellular respiration", "photosynthesis"], "C"),
        (["cell wall", "peptidoglycan", "membrane biogenesis", "envelope",
          "lipopolysaccharide", "capsule"], "M"),
        (["translation", "ribosom"], "J"),
        (["transcription", "rna polymerase", "sigma factor"], "K"),
        (["dna repair", "replication", "recombination"], "L"),
        (["amino acid transport", "amino acid metabolic"], "E"),
        (["signal transduction", "two-component", "chemotaxis"], "T"),
        (["defense", "restriction", "crispr"], "V"),
        (["secretion", "protein transport", "type iii", "type iv"], "U"),
        (["chaperone", "protein folding", "heat shock", "protease"], "O"),
        (["nucleotide metabolic", "purine", "pyrimidine"], "F"),
    ]
    for keywords, cog_letter in mappings:
        if any(kw in go_text for kw in keywords):
            return cog_letter
    return None


def infer_cog_from_product(product_text):
    """
    Infereix la lletra COG a partir del text del producte/funció.
    Fallback local quan les APIs no retornen res.
    """
    if not product_text or str(product_text).strip() in ["", "nan", "-"]:
        return "S"  # Function unknown

    text = str(product_text).lower()

    rules = [
        # Carbohydrates (G)
        (["carbohydrate", "sugar", "glyco", "glucos", "fructos", "lactose",
          "maltose", "sucrose", "galactos", "xylose", "polysaccharide",
          "glycolysis", "gluconeogenesis", "phosphotransferase system", "pts "], "G"),
        # Lipids (I)
        (["lipid", "fatty acid", "acyl", "phospholipid", "membrane lipid",
          "triglyceride", "lipolysis", "beta-oxidation", "lipase"], "I"),
        # Inorganic ions (P)
        (["inorganic ion", "phosphate transport", "sulfate transport", "iron transport",
          "metal ion", "zinc", "copper transport", "potassium transport",
          "abc transporter", "atp-binding cassette", "magnesium", "manganese",
          "cobalt", "ferric", "ferrous", "molybdate", "sulfur"], "P"),
        # Energy (C)
        (["atp synthase", "electron transport", "oxidative phosphorylation",
          "respiratory", "cytochrome", "nadh dehydrogenase", "succinate dehydrogenase",
          "fumarate reductase", "photosystem", "photosynthesis", "energy production",
          "f0f1", "atpase", "proton pump", "quinol", "ubiquinol", "menaquinone",
          "ferredoxin", "hydrogenase", "nitrogenase"], "C"),
        # Cell wall/membrane (M)
        (["cell wall", "peptidoglycan", "murein", "muropeptide", "membrane biogenesis",
          "outer membrane", "lipopolysaccharide", "lps", "capsule biosynthesis",
          "envelope", "mur", "omp", "porin", "penicillin-binding", "pbp",
          "d-ala", "transpeptidase", "transglycosylase", "lytic transglycosylase"], "M"),
        # Translation (J)
        (["ribosom", "translation", "trna", "aminoacyl", "initiation factor",
          "elongation factor", "termination factor", "rrna"], "J"),
        # Transcription (K)
        (["transcription", "rna polymerase", "sigma factor", "regulator",
          "repressor", "activator", "dna-binding", "helix-turn-helix"], "K"),
        # DNA repair/replication (L)
        (["dna repair", "dna replication", "recombinase", "transposase",
          "integrase", "helicase", "dnaa", "dnab", "dnac", "dnag", "ligase",
          "topoisomerase", "gyrase", "nuclease", "exonuclease"], "L"),
        # Amino acid metabolism (E)
        (["amino acid", "glutamate", "glutamine", "aspartate", "asparagine",
          "serine", "threonine", "cysteine", "methionine", "lysine",
          "arginine", "ornithine", "proline", "tryptophan", "phenylalanine",
          "tyrosine", "histidine", "valine", "leucine", "isoleucine"], "E"),
        # Coenzyme metabolism (H)
        (["coenzyme", "cofactor", "biotin", "folate", "riboflavin", "thiamine",
          "pantothenate", "lipoate", "cobalamin", "vitamin", "nad synthesis",
          "fad", "nad biosynthesis"], "H"),
        # Signal transduction (T)
        (["signal transduction", "two-component", "sensor histidine kinase",
          "response regulator", "chemotaxis", "second messenger", "camp",
          "cgmp", "di-gmp", "diguanylate"], "T"),
        # Motility (N)
        (["flagell", "motility", "chemotaxis protein", "fli", "flh", "flg",
          "mot", "fla"], "N"),
        # Secretion (U)
        (["secretion", "type iii secretion", "type iv secretion", "type vi",
          "type ii secretion", "sec translocon", "tat pathway", "twin-arginine"], "U"),
        # Defense (V)
        (["defense", "restriction", "modification", "crispr", "cas protein",
          "bacteriocin", "antibiotic resistance", "efflux pump"], "V"),
        # Chaperones (O)
        (["chaperone", "heat shock protein", "hsp", "groel", "groes", "dnak",
          "dnaj", "grpe", "lon protease", "clp protease", "protease"], "O"),
        # Nucleotide metabolism (F)
        (["nucleotide", "purine", "pyrimidine", "guanosine", "adenosine",
          "thymidine", "uridine", "cytidine", "phosphoribosyl"], "F"),
        # Mobilome (X)
        (["transposon", "insertion sequence", "is element", "prophage",
          "phage", "integron", "plasmid"], "X"),
        # Unknown (S)
        (["hypothetical protein", "unknown function", "uncharacterized",
          "putative", "conserved hypothetical"], "S"),
        # General function (R)
        (["predicted", "possible", "probable"], "R"),
    ]

    for keywords, cog_letter in rules:
        if any(kw in text for kw in keywords):
            return cog_letter

    return "S"  # Per defecte: Function unknown


def get_cog_category(locus_tag, gene_name, product, use_api=True):
    """
    Estratègia multi-font per obtenir la categoria COG:
    1. UniProt (API)
    2. NCBI Gene (API)
    3. Inferència local per paraules clau del producte
    """
    cog_letter = None

    if use_api:
        # Intent 1: UniProt
        cog_letter = query_uniprot_cog(gene_name, locus_tag)
        if cog_letter and len(cog_letter) == 1:
            return cog_letter, "UniProt"

        # Pausa per no sobrecarregar les APIs
        time.sleep(0.3)

        # Intent 2: NCBI
        cog_letter = query_eggnog_by_locus(locus_tag)
        if cog_letter and len(cog_letter) == 1:
            return cog_letter, "NCBI"

        time.sleep(0.2)

    # Intent 3: inferència local
    cog_letter = infer_cog_from_product(product)
    return cog_letter, "Keyword inference"


# ─────────────────────────────────────────────
# PROCESSAMENT PRINCIPAL
# ─────────────────────────────────────────────

def process_genes(input_file, output_file, use_api=True):
    print(f"\n📂 Llegint fitxer: {input_file}")
    df = pd.read_excel(input_file)

    print(f"   Columnes detectades: {list(df.columns)}")
    print(f"   Gens trobats: {len(df)}\n")

    # Detectar noms de columnes automàticament si no coincideixen
    col_map = auto_detect_columns(df)

    results = []
    print("🔍 Classificant gens per categoria COG...")
    for i, row in df.iterrows():
        locus  = str(row.get(col_map["locus"], "")).strip()
        gene   = str(row.get(col_map["gene"], "")).strip()
        product = str(row.get(col_map["product"], "")).strip()
        location = str(row.get(col_map["location"], "")).strip()
        strand = str(row.get(col_map["strand"], "")).strip()
        reg    = str(row.get(col_map["regulation"], "")).strip()

        print(f"   [{i+1}/{len(df)}] {locus} ({gene})...", end=" ")
        cog_letter, source = get_cog_category(locus, gene, product, use_api=use_api)
        cog_name = COG_CATEGORIES.get(cog_letter, "Function unknown")
        print(f"→ {cog_letter}: {cog_name} [{source}]")

        results.append({
            "Locus Tag":          locus,
            "Gene Name":          gene,
            "Product/Function":   product,
            "Location":           location,
            "Strand":             strand,
            "Regulation":         reg,
            "COG Letter":         cog_letter,
            "COG Category":       cog_name,
            "Source":             source,
        })

    df_all = pd.DataFrame(results)

    # Filtrar les 4 categories d'interès
    df_filtered = df_all[df_all["COG Category"].isin(TARGET_CATEGORIES)].copy()
    df_filtered["Target Category"] = df_filtered["COG Category"].map(CATEGORY_DISPLAY)

    print(f"\n✅ Gens classificats: {len(df_all)}")
    print(f"   Gens en les 4 categories d'interès: {len(df_filtered)}")

    write_excel(df_all, df_filtered, output_file)
    print(f"\n💾 Fitxer generat: {output_file}")


def auto_detect_columns(df):
    """
    Detecta automàticament les columnes del DataFrame
    per noms aproximats (insensible a majúscules).
    """
    cols = {c.lower(): c for c in df.columns}

    def find(patterns):
        for p in patterns:
            for k, v in cols.items():
                if p in k:
                    return v
        return df.columns[0]  # fallback

    return {
        "locus":      find(["locus", "tag", "id"]),
        "gene":       find(["gene name", "gene", "name"]),
        "product":    find(["product", "function", "description", "annotation"]),
        "location":   find(["location", "position", "coord", "start", "genome"]),
        "strand":     find(["strand", "chain", "direction", "sense"]),
        "regulation": find(["regulation", "express", "differenti", "up", "down", "over", "repres"]),
    }


# ─────────────────────────────────────────────
# ESCRIPTURA EXCEL
# ─────────────────────────────────────────────

def write_excel(df_all, df_filtered, output_file):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="All Genes - COG Classification", index=False)
        df_filtered.to_excel(writer, sheet_name="Target Categories", index=False)

    wb = load_workbook(output_file)
    format_sheet_all(wb["All Genes - COG Classification"], df_all)
    format_sheet_filtered(wb["Target Categories"], df_filtered)
    wb.save(output_file)


HEADER_FILL  = PatternFill("solid", start_color="2F4F8F")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BODY_FONT    = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

UP_FILL   = PatternFill("solid", start_color="C6EFCE")  # verd per sobreexpresat
DOWN_FILL = PatternFill("solid", start_color="FFC7CE")  # vermell per reprimit


def style_header(sheet):
    for cell in sheet[1]:
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    sheet.row_dimensions[1].height = 30


def style_body_row(row_cells, reg_value="", bg_fill=None):
    reg_lower = str(reg_value).lower()
    for cell in row_cells:
        cell.font      = BODY_FONT
        cell.border    = THIN_BORDER
        cell.alignment = LEFT_ALIGN
        if bg_fill:
            cell.fill = bg_fill
        elif "over" in reg_lower or "up" in reg_lower or "sobre" in reg_lower:
            cell.fill = UP_FILL
        elif "repres" in reg_lower or "down" in reg_lower or "reprimit" in reg_lower:
            cell.fill = DOWN_FILL


def set_col_widths(sheet, widths):
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width


def format_sheet_all(sheet, df):
    style_header(sheet)
    reg_col_idx = None
    for i, col in enumerate(df.columns, 1):
        if "regulation" in col.lower():
            reg_col_idx = i
            break

    for row in sheet.iter_rows(min_row=2):
        reg_val = ""
        if reg_col_idx:
            reg_val = str(row[reg_col_idx - 1].value or "")
        style_body_row(row, reg_value=reg_val)

    set_col_widths(sheet, {
        "A": 16, "B": 14, "C": 40, "D": 20, "E": 8,
        "F": 15, "G": 12, "H": 42, "I": 20,
    })
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def format_sheet_filtered(sheet, df):
    style_header(sheet)

    # Columna de Target Category
    target_col_idx = None
    for i, col in enumerate(df.columns, 1):
        if "target" in col.lower():
            target_col_idx = i
            break
    reg_col_idx = None
    for i, col in enumerate(df.columns, 1):
        if "regulation" in col.lower():
            reg_col_idx = i
            break

    for row in sheet.iter_rows(min_row=2):
        cat_val = ""
        reg_val = ""
        if target_col_idx:
            cat_val = str(row[target_col_idx - 1].value or "")
        if reg_col_idx:
            reg_val = str(row[reg_col_idx - 1].value or "")

        # Color de fons per categoria
        bg = None
        for cat_key, color in CATEGORY_COLORS.items():
            if cat_key in cat_val:
                bg = PatternFill("solid", start_color=color)
                break

        style_body_row(row, reg_value=reg_val, bg_fill=bg)

    set_col_widths(sheet, {
        "A": 16, "B": 14, "C": 40, "D": 20, "E": 8,
        "F": 15, "G": 12, "H": 42, "I": 20, "J": 45,
    })
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


# ─────────────────────────────────────────────
# PUNT D'ENTRADA
# ─────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("""
Ús:  python classify_genes.py <fitxer_entrada.xlsx> <fitxer_sortida.xlsx> [--no-api]

  --no-api    Salta les consultes a UniProt/NCBI i utilitza únicament
              inferència per paraules clau (més ràpid, menys precís)

Exemple:
  python classify_genes.py gens_DEG.xlsx resultats_COG.xlsx
  python classify_genes.py gens_DEG.xlsx resultats_COG.xlsx --no-api
""")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2]
    use_api     = "--no-api" not in sys.argv

    if not use_api:
        print("⚡ Mode ràpid: sense consultes API (inferència per paraules clau)")

    process_genes(input_file, output_file, use_api=use_api)
